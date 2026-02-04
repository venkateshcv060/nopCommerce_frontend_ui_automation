import openpyxl


import openpyxl


def get_excel_data(file_path, sheet_name):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    headers = []

    for col in range(1, sheet.max_column + 1):
        headers.append(sheet.cell(row=1, column=col).value)


    for row in range(2, sheet.max_row + 1):
        row_dict = {}

        for col in range(1, sheet.max_column + 1):

            key = headers[col - 1]
            value = sheet.cell(row=row, column=col).value

            row_dict[key] = value

        data.append(row_dict)

    return data
